"""AMFI half-yearly "Average Market Capitalization of listed companies" fetcher.

FROZEN SPEC: docs/superpowers/specs/2026-07-28-amfi-band-crossing-design.md
This script implements the spec's GATING RECON STEP. It is deliverable 1 of 2;
kite/research/amfi_band_study.py consumes what this writes.

--------------------------------------------------------------------------
RECON GATE (the spec's halt condition, implemented literally)
--------------------------------------------------------------------------
The spec says: "builder must first verify these historical lists are
downloadable from amfiindia.com ... If lists cannot be retrieved for >=12 of
the ~17 reviews, the study HALTS as data-blocked -- that is a recorded
outcome, not a failure to route around."

So this script's FIRST action is to discover + count retrievable reviews and
either print RECON: PASS or print RECON: DATA-BLOCKED and exit non-zero
WITHOUT writing a partial dataset. There is deliberately no fallback path
that recomputes market caps from the price panel -- the spec forbids it
("We use AMFI's published ranks, NOT a recomputation ... our panel lacks
shares-outstanding to recompute honestly").

--------------------------------------------------------------------------
WHAT RECON ACTUALLY FOUND (2026-07-28) -- three surprises, all documented
--------------------------------------------------------------------------
(1) The URL suggested in the build order,
    /research-information/other-data/categorization-of-stocks, is a 404.
    amfiindia.com was rebuilt as a Next.js/Strapi site; the live page is
    https://www.amfiindia.com/otherdata/categorisation-of-stocks
    and the nav is client-rendered, so link discovery scrapes that page's
    server-rendered HTML for /Themes/Theme1/downloads/ + portal /spages/
    hrefs rather than walking the menu.

(2) The archive is COMPLETE and then some: 18 half-yearly lists, Jul-Dec 2017
    through Jan-Jun 2026, each as both .xlsx and .pdf. 18 lists => 17
    consecutive-pair transitions, exactly the "~17 reviews" the spec sized
    for. The gate passes with zero margin needed.

(3) AMFI does NOT key these lists by company name alone, contrary to the
    build order's assumption. Every file carries, per row:
        Sr. No. | Company name | ISIN | BSE Symbol | BSE 6m avg mcap |
        NSE Symbol | NSE 6m avg mcap | MSEI Symbol
    So there is an explicit NSE Symbol column and an explicit ISIN. No fuzzy
    name->symbol matching is required or performed. See the JOIN section
    below for what this means and what is still reported for inspection.

    What AMFI does NOT publish in these files is a Large/Mid/Small CATEGORY
    column. It publishes the RANK (Sr. No.). Category is therefore derived
    from AMFI's own published rank via SEBI's frozen boundaries
    (1-100 large, 101-250 mid, 251+ small) -- which is what the spec
    describes ("keyed to AMFI's half-yearly market-cap ranking (top 100 =
    large, 101-250 = mid, 251+ = small)"). This is NOT a recomputation: the
    rank is taken verbatim from the file, only the label is applied.

--------------------------------------------------------------------------
RANK SEMANTICS (verified, not assumed)
--------------------------------------------------------------------------
Sr. No. is AMFI's published rank over ALL listed companies (~5,100), sorted
descending by the AVERAGE of the exchange-wise 6-month average full market
caps -- not by NSE alone. Verified on the Jan-Jun 2018 list: ONGC (NSE
234,525.82) is ranked 9 and SBIN (NSE 232,532.14) is ranked 10 even though
SBIN's BSE figure (236,152.86) is the larger of the four numbers; the
ordering only reproduces if you average BSE and NSE per company
(ONGC 234,474.33 > SBIN 234,342.50). The script recomputes that average into
an `avg_mcap_cr` column for INSPECTION ONLY and reports how often AMFI's own
ordering disagrees with it -- it never re-ranks. Ranks are used verbatim.

Because the rank spans all listed companies including BSE-only ones, the
top-100 / top-250 cut lines are AMFI's real cut lines, not an NSE-only
approximation. The spec anticipated this ("AMFI ranks use NSE+BSE combined
mcap; borderline misclassification vs our NSE-only panel affects joins, not
ranks (we use their ranks verbatim)").

--------------------------------------------------------------------------
JOIN (reviewer-inspection material, per the build order)
--------------------------------------------------------------------------
Two different joins are needed and they use different keys on purpose:

  a) LIST-TO-LIST (detecting a band crossing between consecutive reviews):
     joined on ISIN. ISIN is stable across ticker renames and is present in
     every file; symbol is not stable (renames, merges). amfi_band_study.py
     does this join.

  b) LIST-TO-PANEL (getting prices): joined on the file's own NSE Symbol
     against the bhavcopy EQ symbol set. This is an EXACT string join. There
     is no fuzzy matching anywhere in this pipeline, so there are no fuzzy
     matches to hide.

     Rows with no NSE Symbol ('' or '-') are BSE-only listings. They are not
     tradeable on our NSE panel and are reported as such rather than being
     force-matched by name. A name->symbol fallback was considered and
     rejected: the repo has no NSE symbol->company-name master to match
     against (bhavcopy carries SYMBOL only, no name), so any such mapping
     would be invented rather than derived.

This script reports, per review: rows, rows with an NSE symbol, and the exact
join rate against the panel's EQ symbols on the trading day nearest that
list's publication date -- reported BOTH over all ~5,100 rows and over the
top 300 ranks, because only the top 300 can possibly produce a band crossing
and a 60% overall join rate driven by BSE-only microcaps says nothing about
whether the tradeable population joined. Every unjoined top-300 row is
printed by name so the reviewer can eyeball it.

--------------------------------------------------------------------------
PUBLICATION DATES (spec: "obtain actual dates; if unverifiable, use the 5th
trading day of Jan/Jul and say so per event")
--------------------------------------------------------------------------
The categorisation page carries no publication-date text. The only server-
side signal is each file's HTTP Last-Modified. That signal is CONTAMINATED
for the historical archive: every file from Jan-Jun 2025 and earlier returns
"19 Aug 2025", a bulk re-upload timestamp from the site migration, not a
publication date.

So this script accepts Last-Modified as a VERIFIED publication date only when
it falls inside the month the list was actually due to be published (January
for a Jul-Dec list, July for a Jan-Jun list, of the correct year); otherwise
it records pub_date_source='fallback' and leaves the date for the study
script to fill with the spec's 5th-trading-day rule. As of 2026-07-28 that
verifies exactly 2 of 18 (Jul-Dec 2025 -> 2026-01-05, Jan-Jun 2026 ->
2026-07-07) and flags the other 16 as fallback. Both are carried per-review
in the manifest so every event can state which it used.

--------------------------------------------------------------------------
OUTPUTS
--------------------------------------------------------------------------
data/amfi_bands/raw/<original filename>          untouched downloaded .xlsx
data/amfi_bands/amfi_<YYYY>_H<1|2>.csv           normalized, one per review
data/amfi_bands/amfi_reviews.csv                 manifest (one row/review)

Normalized CSV columns:
    rank, company_name, isin, bse_symbol, bse_avg_mcap_cr, nse_symbol,
    nse_avg_mcap_cr, avg_mcap_cr, category
`category` is derived from `rank` (1-100 LARGE, 101-250 MID, 251+ SMALL).
`avg_mcap_cr` is the BSE/NSE mean, inspection only, never used to rank.

Period labelling is by the DATA period, not the publication date:
amfi_2025_H2.csv holds the Jul-Dec 2025 list, which was published Jan 2026.

Usage:
    python kite/research/fetch_amfi_bands.py                # recon + download
    python kite/research/fetch_amfi_bands.py --recon-only   # gate check only
    python kite/research/fetch_amfi_bands.py --refresh      # re-download raw
"""

import argparse
import re
import sys
import time
from pathlib import Path

import pandas as pd
import requests

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / 'data' / 'amfi_bands'
RAW_DIR = OUT_DIR / 'raw'
PANEL_DIR = ROOT / 'data' / 'bhavcopy_full'

INDEX_URL = 'https://www.amfiindia.com/otherdata/categorisation-of-stocks'
HEADERS = {
    'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                   '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'),
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Referer': INDEX_URL,
}

# Spec's gate: >=12 of ~17 reviews retrievable, else DATA-BLOCKED.
MIN_REVIEWS_REQUIRED = 12

# SEBI / AMFI frozen category boundaries on AMFI's published rank.
LARGE_MAX_RANK = 100
MID_MAX_RANK = 250

MONTHS = {
    'jan': 1, 'january': 1, 'feb': 2, 'mar': 3, 'march': 3, 'apr': 4,
    'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7, 'aug': 8,
    'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12, 'december': 12,
}

_OUT = []


def log(msg=''):
    print(msg, flush=True)
    _OUT.append(str(msg))


def flush_log():
    """Persist the run log even on a DATA-BLOCKED exit -- the spec calls that
    a RECORDED OUTCOME, so it has to leave a record on disk."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / 'fetch_amfi_bands.log').write_text('\n'.join(_OUT) + '\n', encoding='utf-8')


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------
def period_from_filename(name):
    """Map an AMFI download filename to (year, half) of the DATA period.

    AMFI has used at least four naming styles across 2017-2026; all are
    handled explicitly rather than by a single clever regex, because a silent
    mis-parse here would silently mis-date every event downstream:

        Avg. Market Capitalization ... during -Jul-Dec 2017.xlsx
        Average Market Capitalization ... during Jan - Jun 2020_Final.xlsx
        Average Market Capitalisation ... during Jul - Dec 2021.xlsx
        AverageMarketCapitalization...sixmonthsended31Dec2024.xlsx
        AverageMarketCapitalization30Jun2026.xlsx

    Returns (year, half) or None if unrecognised (caller reports it).
    """
    n = re.sub(r'%20', ' ', name)
    n = re.sub(r'[_]+', ' ', n).strip()

    # Style A: "...ended 31Dec2024" / "30Jun2026" / "30June2022" -- an END date.
    m = re.search(r'(\d{1,2})\s*([A-Za-z]{3,9})\s*(\d{4})', n)
    if m:
        mon = MONTHS.get(m.group(2).lower())
        year = int(m.group(3))
        if mon == 12:
            return year, 2
        if mon == 6:
            return year, 1

    # Style B: an explicit month RANGE "Jan-June 2021", "Jul - Dec 2021".
    m = re.search(r'([A-Za-z]{3,9})\s*-\s*([A-Za-z]{3,9})\s*(\d{4})', n)
    if m:
        m1 = MONTHS.get(m.group(1).lower())
        m2 = MONTHS.get(m.group(2).lower())
        year = int(m.group(3))
        if m1 == 1 and m2 == 6:
            return year, 1
        if m1 == 7 and m2 == 12:
            return year, 2

    return None


def discover_reviews(session):
    """Scrape the categorisation page for per-review download links."""
    log(f'Fetching index: {INDEX_URL}')
    r = session.get(INDEX_URL, headers=HEADERS, timeout=60)
    log(f'  HTTP {r.status_code}, {len(r.text)} bytes')
    if r.status_code != 200:
        return {}, f'index page returned HTTP {r.status_code}'

    hrefs = re.findall(r'href=["\']([^"\']+)["\']', r.text)
    # Trailing/leading whitespace inside href values is present in the live
    # HTML for at least one 2020 link -- strip before use or the GET 404s.
    hrefs = [h.strip() for h in dict.fromkeys(hrefs)
             if re.search(r'(Themes/Theme1/downloads|/spages/|/uploads/)', h)]

    reviews = {}
    unparsed = []
    for h in hrefs:
        url = h if h.startswith('http') else ('https://www.amfiindia.com' + h)
        # MUST unquote before any content test: the archive mixes literal
        # spaces, %20 and underscores as word separators in the SAME page, so
        # matching on the raw href silently drops whole eras (this bit once:
        # a raw-href filter found 10 of the 18 lists and tripped the gate).
        fname = requests.utils.unquote(url.split('/')[-1]).strip()
        norm = re.sub(r'[_\-]+', ' ', fname)
        if not re.search(r'(?i)market\s*capitali[sz]', norm):
            continue
        ext = fname.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'pdf'):
            continue
        per = period_from_filename(fname)
        if per is None:
            unparsed.append(fname)
            continue
        slot = reviews.setdefault(per, {})
        # Prefer xlsx; keep pdf only as a recorded alternate (not parsed).
        if ext in ('xlsx', 'xls'):
            slot.setdefault('xlsx_url', url)
            slot.setdefault('xlsx_name', fname)
        else:
            slot.setdefault('pdf_url', url)

    if unparsed:
        log(f'  NOTE: {len(unparsed)} download link(s) had an unrecognised '
            f'period pattern and were skipped: {unparsed}')
    return reviews, None


# ---------------------------------------------------------------------------
# Recon gate
# ---------------------------------------------------------------------------
def expected_pub_month_year(year, half):
    """The month/year AMFI was due to publish the list for (year, half)."""
    return (7, year) if half == 1 else (1, year + 1)


def probe_and_gate(session, reviews):
    """HEAD every candidate xlsx; count what is really retrievable.

    Also harvests Last-Modified as a publication-date candidate (see module
    docstring for why it is only conditionally trusted).
    """
    log('')
    log('--- RECON: probing each review file (HEAD) ---')
    ok, bad = [], []
    for per in sorted(reviews):
        year, half = per
        slot = reviews[per]
        url = slot.get('xlsx_url')
        label = f'{year} H{half}'
        if not url:
            log(f'  {label}: NO xlsx link (pdf only) -> not retrievable')
            bad.append(per)
            continue
        try:
            h = session.head(url, headers=HEADERS, timeout=60, allow_redirects=True)
        except Exception as e:
            log(f'  {label}: HEAD failed {type(e).__name__}: {e}')
            bad.append(per)
            continue
        size = h.headers.get('content-length')
        lastmod = h.headers.get('last-modified')
        if h.status_code != 200:
            log(f'  {label}: HTTP {h.status_code} -> not retrievable')
            bad.append(per)
            continue

        pub_date, pub_src = None, 'fallback'
        if lastmod:
            try:
                lm = pd.Timestamp(lastmod).tz_convert(None)
            except Exception:
                lm = None
            if lm is not None:
                slot['last_modified'] = lm.date().isoformat()
                exp_m, exp_y = expected_pub_month_year(year, half)
                if lm.month == exp_m and lm.year == exp_y:
                    pub_date, pub_src = lm.normalize().date().isoformat(), 'last-modified'
        slot['pub_date'] = pub_date
        slot['pub_date_source'] = pub_src
        log(f'  {label}: HTTP 200, {size} bytes, Last-Modified={lastmod} '
            f'-> pub_date={pub_date or "(fallback)"} [{pub_src}]')
        ok.append(per)

    n_ok = len(ok)
    n_pairs = max(n_ok - 1, 0)
    log('')
    log(f'RECON RESULT: {n_ok} of {len(reviews)} discovered reviews are retrievable '
        f'({n_pairs} consecutive-pair transitions).')
    log(f'Spec gate: >= {MIN_REVIEWS_REQUIRED} of ~17 reviews required.')
    if n_ok < MIN_REVIEWS_REQUIRED:
        log('')
        log('RECON: DATA-BLOCKED')
        log(f'  Only {n_ok} reviews retrievable, below the frozen gate of '
            f'{MIN_REVIEWS_REQUIRED}. Per the spec this is a RECORDED OUTCOME, '
            f'not a routing-around problem: the study HALTS. Do NOT substitute '
            f'a market-cap recomputation from the price panel -- the spec '
            f'forbids it (no shares-outstanding data).')
        return ok, False
    log('')
    log('RECON: PASS -- proceeding to download and normalize.')
    return ok, True


# ---------------------------------------------------------------------------
# Download + parse
# ---------------------------------------------------------------------------
def download(session, url, dest, refresh=False):
    if dest.exists() and not refresh and dest.stat().st_size > 0:
        return dest, 'cached'
    for attempt in range(3):
        try:
            r = session.get(url, headers=HEADERS, timeout=180)
            if r.status_code == 200 and r.content[:2] == b'PK':
                dest.write_bytes(r.content)
                return dest, 'downloaded'
            last = f'HTTP {r.status_code}, magic={r.content[:4]!r}'
        except Exception as e:
            last = f'{type(e).__name__}: {e}'
        time.sleep(2 * (attempt + 1))
    return None, last


COLSPECS = [
    ('rank', r'sr\.?\s*no'),
    ('company_name', r'company\s*name'),
    ('isin', r'isin'),
    ('bse_symbol', r'bse\s*symbol'),
    ('bse_avg_mcap_cr', r'bse.*(?:market\s*cap|mkt\s*cap)'),
    ('nse_symbol', r'nse\s*symbol'),
    ('nse_avg_mcap_cr', r'nse.*(?:market\s*cap|mkt\s*cap)'),
]


def parse_xlsx(path):
    """Parse one AMFI xlsx into a normalized DataFrame.

    Header position and sheet name both vary across the archive (sheet names
    seen: 'Avg-6M-June 2018 BSE NSE MSE', 'AVG-6M-JAN2021' -- which is
    mislabelled, that file is Jul-Dec 2021 -- 'Sheet1', 'FINAL'). So: always
    take the first sheet, and locate the header by CONTENT (a row containing
    both a 'Sr. No.'-ish and a 'Company name'-ish cell) within the first 10
    rows. Never trust sheet names or a fixed row offset.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(['' if c is None else str(c).strip() for c in row])
    wb.close()

    hdr_i, colmap = None, None
    for i, row in enumerate(rows[:10]):
        low = [c.lower() for c in row]
        if any(re.search(r'sr\.?\s*no', c) for c in low) and \
           any(re.search(r'company\s*name', c) for c in low):
            cm = {}
            for key, pat in COLSPECS:
                for j, c in enumerate(low):
                    if re.search(pat, c):
                        cm[key] = j
                        break
            hdr_i, colmap = i, cm
            break
    if hdr_i is None:
        raise ValueError(f'{path.name}: could not locate a header row in the first 10 rows')
    missing = [k for k, _ in COLSPECS if k not in colmap]
    if missing:
        raise ValueError(f'{path.name}: header found but missing columns {missing} '
                         f'(header row was: {rows[hdr_i]!r})')

    recs = []
    for row in rows[hdr_i + 1:]:
        def get(k):
            j = colmap[k]
            return row[j] if j < len(row) else ''
        rk = get('rank')
        if not re.fullmatch(r'\d+(\.0)?', rk or ''):
            continue
        recs.append({
            'rank': int(float(rk)),
            'company_name': get('company_name'),
            'isin': get('isin').upper(),
            'bse_symbol': get('bse_symbol'),
            'bse_avg_mcap_cr': get('bse_avg_mcap_cr'),
            'nse_symbol': get('nse_symbol'),
            'nse_avg_mcap_cr': get('nse_avg_mcap_cr'),
        })
    df = pd.DataFrame(recs)
    if df.empty:
        raise ValueError(f'{path.name}: header parsed but zero data rows matched')

    # AMFI writes '-' (and sometimes blank) for "not listed on this exchange".
    for c in ('bse_symbol', 'nse_symbol'):
        df[c] = df[c].replace({'-': '', 'NA': '', 'na': ''}).str.strip()
    for c in ('bse_avg_mcap_cr', 'nse_avg_mcap_cr'):
        df[c] = pd.to_numeric(df[c].replace({'-': ''}), errors='coerce')

    # Inspection-only combined figure (mean of the exchange-wise averages).
    # NEVER used to rank -- AMFI's published rank is authoritative.
    df['avg_mcap_cr'] = df[['bse_avg_mcap_cr', 'nse_avg_mcap_cr']].mean(axis=1, skipna=True)
    df['category'] = pd.cut(
        df['rank'], bins=[0, LARGE_MAX_RANK, MID_MAX_RANK, 10 ** 9],
        labels=['LARGE', 'MID', 'SMALL']).astype(str)
    return df


def validate(df, label):
    """Report (do not fix) structural problems a reviewer would want to see."""
    notes = []
    n = len(df)
    if list(df['rank']) != list(range(1, n + 1)):
        dupes = int(df['rank'].duplicated().sum())
        notes.append(f'rank is NOT a contiguous 1..{n} sequence '
                     f'(min={df["rank"].min()}, max={df["rank"].max()}, dupes={dupes})')
    v = df.dropna(subset=['avg_mcap_cr']).sort_values('rank')['avg_mcap_cr'].to_numpy()
    inversions = int((v[1:] > v[:-1]).sum()) if len(v) > 1 else 0
    if inversions:
        notes.append(f'{inversions} rank/avg-mcap inversions vs the recomputed '
                     f'BSE/NSE mean (expected: AMFI rounds and may use more '
                     f'exchanges; rank still used verbatim)')
    n_isin = int((df['isin'].str.len() == 12).sum())
    if n_isin < n:
        notes.append(f'{n - n_isin} row(s) lack a well-formed 12-char ISIN')
    for note in notes:
        log(f'    CHECK {label}: {note}')
    return {'rank_contiguous': list(df['rank']) == list(range(1, n + 1)),
            'mcap_inversions': inversions, 'isin_ok': n_isin}


# ---------------------------------------------------------------------------
# Join-rate reporting against the price panel
# ---------------------------------------------------------------------------
_PANEL_CACHE = {}


def panel_dates():
    if 'dates' not in _PANEL_CACHE:
        rx = re.compile(r'sec_bhavdata_full_(\d{2})(\d{2})(\d{4})\.csv$', re.I)
        out = {}
        for f in PANEL_DIR.glob('sec_bhavdata_full_*.csv'):
            m = rx.search(f.name)
            if m:
                out[pd.Timestamp(year=int(m.group(3)), month=int(m.group(2)),
                                 day=int(m.group(1)))] = f
        _PANEL_CACHE['dates'] = dict(sorted(out.items()))
    return _PANEL_CACHE['dates']


def panel_symbols_near(target):
    """EQ symbol set on the panel trading day nearest `target` (or None)."""
    d = panel_dates()
    if not d:
        return None, None
    keys = list(d)
    nearest = min(keys, key=lambda k: abs((k - target).days))
    if abs((nearest - target).days) > 45:
        return None, None
    if nearest in _PANEL_CACHE:
        return _PANEL_CACHE[nearest], nearest
    df = pd.read_csv(d[nearest], dtype=str)
    df.columns = df.columns.str.strip()
    syms = set(df.loc[df['SERIES'].str.strip() == 'EQ', 'SYMBOL'].str.strip())
    _PANEL_CACHE[nearest] = syms
    return syms, nearest


def join_report(df, year, half):
    """Exact NSE-symbol join rate vs the panel, overall and for the top 300."""
    exp_m, exp_y = expected_pub_month_year(year, half)
    target = pd.Timestamp(year=exp_y, month=exp_m, day=7)
    syms, nearest = panel_symbols_near(target)
    n = len(df)
    n_nse = int((df['nse_symbol'] != '').sum())
    if syms is None:
        log(f'    JOIN {year} H{half}: {n_nse}/{n} rows carry an NSE symbol; '
            f'panel has no data near {target.date()} -> join rate N/A '
            f'(review predates the price panel)')
        return {'n_rows': n, 'n_nse': n_nse, 'n_join': -1, 'n_top300': -1,
                'n_top300_join': -1, 'panel_day': ''}

    joined = df['nse_symbol'].isin(syms) & (df['nse_symbol'] != '')
    n_join = int(joined.sum())
    top = df[df['rank'] <= 300]
    t_join = int((top['nse_symbol'].isin(syms) & (top['nse_symbol'] != '')).sum())
    log(f'    JOIN {year} H{half}: {n_nse}/{n} rows have an NSE symbol; '
        f'{n_join}/{n} ({100 * n_join / n:.1f}%) join the panel EQ set of '
        f'{nearest.date()}; TOP-300: {t_join}/{len(top)} '
        f'({100 * t_join / max(len(top), 1):.1f}%)')
    miss = top[~(top['nse_symbol'].isin(syms) & (top['nse_symbol'] != ''))]
    if len(miss):
        log(f'      top-300 rows that did NOT join ({len(miss)}) -- '
            f'inspect these, no fuzzy fallback was applied:')
        for _, r in miss.iterrows():
            log(f'        rank {int(r["rank"]):>3}  {r["company_name"][:44]:<44} '
                f'nse={r["nse_symbol"] or "(none)":<12} bse={r["bse_symbol"] or "(none)"}')
    return {'n_rows': n, 'n_nse': n_nse, 'n_join': n_join, 'n_top300': len(top),
            'n_top300_join': t_join, 'panel_day': nearest.date().isoformat()}


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--recon-only', action='store_true',
                    help='run the gate check and stop (writes nothing)')
    ap.add_argument('--refresh', action='store_true',
                    help='re-download raw files even if already cached')
    args = ap.parse_args()

    log('=' * 78)
    log('AMFI BAND LIST FETCHER -- gating recon per frozen spec')
    log('  spec: docs/superpowers/specs/2026-07-28-amfi-band-crossing-design.md')
    log(f'  gate: >= {MIN_REVIEWS_REQUIRED} of ~17 reviews retrievable, else DATA-BLOCKED')
    log('  ranks: AMFI published Sr. No. used VERBATIM; no recomputation')
    log(f'  category: rank 1-{LARGE_MAX_RANK} LARGE, {LARGE_MAX_RANK + 1}-{MID_MAX_RANK} MID, '
        f'{MID_MAX_RANK + 1}+ SMALL')
    log('=' * 78)

    session = requests.Session()
    reviews, err = discover_reviews(session)
    if err:
        log('')
        log('RECON: DATA-BLOCKED')
        log(f'  Could not read the AMFI index page: {err}')
        flush_log()
        sys.exit(2)

    log(f'Discovered {len(reviews)} half-yearly review(s) on the index page:')
    for (y, h) in sorted(reviews):
        s = reviews[(y, h)]
        log(f'  {y} H{h}  xlsx={"yes" if s.get("xlsx_url") else "NO"}  '
            f'pdf={"yes" if s.get("pdf_url") else "no"}  {s.get("xlsx_name", "")}')

    ok, passed = probe_and_gate(session, reviews)
    if not passed:
        flush_log()
        sys.exit(3)
    if args.recon_only:
        log('')
        log('--recon-only: stopping here, nothing written.')
        return

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    log('')
    log('--- DOWNLOAD + NORMALIZE + JOIN REPORT ---')
    manifest = []
    for per in ok:
        year, half = per
        slot = reviews[per]
        label = f'{year} H{half}'
        raw = RAW_DIR / slot['xlsx_name']
        path, how = download(session, slot['xlsx_url'], raw, refresh=args.refresh)
        if path is None:
            log(f'  {label}: DOWNLOAD FAILED ({how}) -- skipping')
            continue
        try:
            df = parse_xlsx(path)
        except Exception as e:
            log(f'  {label}: PARSE FAILED {type(e).__name__}: {e} -- skipping')
            continue

        cats = df['category'].value_counts()
        log(f'  {label}: {how}, {len(df)} rows  '
            f'[LARGE={cats.get("LARGE", 0)} MID={cats.get("MID", 0)} SMALL={cats.get("SMALL", 0)}]  '
            f'sheet-source={slot["xlsx_name"]}')
        v = validate(df, label)
        j = join_report(df, year, half)

        out_csv = OUT_DIR / f'amfi_{year}_H{half}.csv'
        df.to_csv(out_csv, index=False)
        exp_m, exp_y = expected_pub_month_year(year, half)
        manifest.append({
            'period': f'{year}H{half}',
            'period_year': year, 'period_half': half,
            'pub_month': exp_m, 'pub_year': exp_y,
            'pub_date': slot.get('pub_date') or '',
            'pub_date_source': slot.get('pub_date_source', 'fallback'),
            'last_modified': slot.get('last_modified', ''),
            'url': slot['xlsx_url'], 'raw_file': raw.name, 'csv_file': out_csv.name,
            'n_rows': len(df), 'rank_contiguous': v['rank_contiguous'],
            'mcap_inversions': v['mcap_inversions'],
            'n_nse_symbol': j['n_nse'], 'n_join_panel': j['n_join'],
            'n_top300': j['n_top300'], 'n_top300_join': j['n_top300_join'],
            'panel_day_used': j['panel_day'],
        })

    if len(manifest) < MIN_REVIEWS_REQUIRED:
        log('')
        log('RECON: DATA-BLOCKED (post-download)')
        log(f'  Only {len(manifest)} review(s) survived download+parse, below the '
            f'frozen gate of {MIN_REVIEWS_REQUIRED}.')
        flush_log()
        sys.exit(3)

    man = pd.DataFrame(manifest).sort_values(['period_year', 'period_half'])
    man_path = OUT_DIR / 'amfi_reviews.csv'
    man.to_csv(man_path, index=False)

    log('')
    log('--- SUMMARY ---')
    log(f'Reviews written : {len(man)}  ({len(man) - 1} consecutive-pair transitions)')
    log(f'Period range    : {man.iloc[0]["period"]} .. {man.iloc[-1]["period"]}')
    n_ver = int((man['pub_date_source'] == 'last-modified').sum())
    log(f'Publication date: {n_ver} verified from Last-Modified, '
        f'{len(man) - n_ver} will use the spec fallback (5th trading day of Jan/Jul)')
    log(f'Manifest        : {man_path}')
    log(f'Normalized CSVs : {OUT_DIR}/amfi_YYYY_H*.csv')
    log(f'Raw xlsx copies : {RAW_DIR}')
    log('')
    log('Join-rate table (exact NSE-symbol match, no fuzzy matching anywhere):')
    log(f'  {"period":<8} {"rows":>5} {"nse_sym":>8} {"joined":>7} {"join%":>7} '
        f'{"top300":>7} {"t300%":>7}  pub_date')
    for _, r in man.iterrows():
        jp = ('n/a' if r['n_join_panel'] < 0
              else f'{100 * r["n_join_panel"] / max(r["n_rows"], 1):.1f}%')
        tp = ('n/a' if r['n_top300_join'] < 0
              else f'{100 * r["n_top300_join"] / max(r["n_top300"], 1):.1f}%')
        pd_str = r['pub_date'] if r['pub_date'] else f'(fallback {r["pub_month"]:02d}/{r["pub_year"]})'
        log(f'  {r["period"]:<8} {r["n_rows"]:>5} {r["n_nse_symbol"]:>8} '
            f'{max(r["n_join_panel"], 0):>7} {jp:>7} {max(r["n_top300_join"], 0):>7} '
            f'{tp:>7}  {pd_str}')

    log('')
    log('RECON: PASS -- data available, study may proceed.')
    (OUT_DIR / 'fetch_amfi_bands.log').write_text('\n'.join(_OUT) + '\n', encoding='utf-8')


if __name__ == '__main__':
    main()
